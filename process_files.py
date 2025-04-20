barWidth = 36

def csv_process(inputPath,outputPath,keywords,labLabels,totals):
	
	outputFile = open(outputPath,'w')
	sep = ',' # for CSV: Comma Separated Values
	
	# Write the date and time of when the extractor was used
	dt = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
	outputFile.write(bar(barWidth,'='))
	outputFile.write(f'\nAT Report Extractor Output\nGenerated at {dt}\n')
	outputFile.write(bar(barWidth,'='))

	# Begin reading input directory
	for fileName in os.listdir(inputPath):
		filePath = inputPath + '\\' + fileName
		inputFile = open(filePath,'r')
		for line in inputFile:
			
			# Record the date of the Advisortrac report
			found = line.find('from')
			if (found != -1):
				dateStr = 'Dates ' + line[found:found+29]
				dateList = dateStr.split()
				totals.append(dateList)
				continue # Nothing else is worthwhile on this line
			
			'''
			These are very specific to our AdvisorTrac Reports and is not 
			designed to work with all Reports!
			'''

			newList = list(filter(lambda str: str != '' and str != 'Students\n', line.split('\t')))

			# Criteria to clean up the list
			if (newList[0] == '<!--Group1 Dif-->'):
				newList.pop(0)
			if (newList[0].find('<!--# vis at beginning of report:') != -1):
				tempList = newList[0].split('>')
				newList[0] = tempList[1]
			if newList[0] not in keywords:
				continue
			if newList[0] in labLabels:
				del newList[1:]
			if (newList[0] == 'Grand Total:'):
				newList[1] = newList[1].replace('<b>','').replace('</b>','')
				newList[3] = newList[3].replace('<b>','').replace('</b>','')
			
			totals.append(newList)
		# end of for-loop
	
	# Write to file
	for element in totals:
		if element[0] in labLabels:
			outputFile.write('\n\n' + sep.join([element[0],'Visits','Hours','Students']))
		elif element[0] == 'Grand Total:':
			outputFile.write('\n\n' + sep.join(['Grand Total','Visits','Hours','Students']))
			outputFile.write('\n' + sep.join(element))
		elif element[0] == 'Dates':
			outputFile.write('\n\n\n' + bar(barWidth))
			outputFile.write('\n' + ' '.join(element))
			outputFile.write('\n' + bar(barWidth))
		else:
			outputFile.write('\n' + sep.join(element))
	# end of for-loop
	
	print(f'\nProcess completed successfully. Output file is at \"{outputPath}\".')
	inputFile.close()
	outputFile.close()
	return True

def excel_process(inputPath,outputPath,keywords,labLabels,totals):
	import openpyxl

	# Create a new workbook
	wb = openpyxl.Workbook()

	# Get the active worksheet
	ws = wb.active
	ws.title = "Data Sheet"
    
	barWidth = 36
	sep = ',' # for CSV: Comma Separated Values
	
	# Write the date and time of when the extractor was used
	dt = datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')
	ws['A1'] = '=' * barWidth
	ws['A2'] = f'\nAT Report Extractor Output\nGenerated at {dt}\n'
	ws['A3'] = '=' * barWidth

	# Begin reading input directory
	for fileName in os.listdir(inputPath):
		filePath = inputPath + '\\' + fileName
		inputFile = open(filePath,'r')
		for line in inputFile:
			
			# Record the date of the Advisortrac report
			found = line.find('from')
			if (found != -1):
				dateStr = 'Dates ' + line[found:found+29]
				dateList = dateStr.split()
				totals.append(dateList)
				continue # Nothing else is worthwhile on this line
			
			'''
			These are very specific to our AdvisorTrac Reports and is not 
			designed to work with all Reports!
			'''

			newList = list(filter(lambda str: str != '' and str != 'Students\n', line.split('\t')))

			# Criteria to clean up the list
			if (newList[0] == '<!--Group1 Dif-->'):
				newList.pop(0)
			if (newList[0].find('<!--# vis at beginning of report:') != -1):
				tempList = newList[0].split('>')
				newList[0] = tempList[1]
			if newList[0] not in keywords:
				continue
			if newList[0] in labLabels:
				del newList[1:]
			if (newList[0] == 'Grand Total:'):
				newList[1] = newList[1].replace('<b>','').replace('</b>','')
				newList[3] = newList[3].replace('<b>','').replace('</b>','')
			
			totals.append(newList)
		# end of for-loop
	
	# Writing information to output file
	for row, element in enumerate(totals, start=7):
		if element[0] in labLabels:
			# Write collumn names to cells
			ws[f'A{row}'] = element[0]
			ws[f'B{row}'] = "Visits"
			ws[f'C{row}'] = "Hours"
			ws[f'D{row}'] = "students"
		elif element[0] == 'Grand Total:':
			# Writes collumn names to cells
			ws[f'A{row}'] = "Grand Total"
			ws[f'B{row}'] = "Visits"
			ws[f'C{row}'] = "Hours"
			ws[f'D{row}'] = "students"

            # Writes values to cells
            data_row = row + 1
            ws.cell(row=data_row, column=1, value=element[0])  # "Grand Total:"
            if len(element) > 1:
                ws.cell(row=data_row, column=2, value=element[1])  # Visits value
            if len(element) > 2:
                ws.cell(row=data_row, column=3, value=element[2])  # Hours value
            if len(element) > 3:
                ws.cell(row=data_row, column=4, value=element[3])  # Students value
				
		elif element[0] == 'Dates':
			# Insert dashes into cell
			ws[f'A{row+3}'] = '-' * barWidth
			ws[f'A{row+4}'] = ' '.join(element)
			ws[f'A{row+5}'] = '-' * barWidth

		else:
			ws[f'A{row+1}'] = sep.join(element)
	# end of for-loop

	#saves file
	filename = "sample_data.xlsx"
	wb.save(filename)
	print(f'\nProcess completed successfully. Output file is at \"{filename}\".')
	return True


if __name__ == '__main__':
    excel_process()
    csv_process()